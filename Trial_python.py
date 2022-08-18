import tkinter as tk
from tkinter import *

top = tk.Tk()
inp = ''
dnn = ''




def getInput():
    global inp 
    global dnn
    inp = inputtxt.get()
    dnn = inputtxt1.get()
    top.destroy()



L1 = Label(top, text="File Name")
L1.pack( side = LEFT)
inputtxt = Entry(top)
inputtxt.pack(side = LEFT)

L2 = Label(top, text = 'Prefix')
L2.pack(side = RIGHT)
inputtxt1 = Entry(top)
inputtxt1.pack(side = RIGHT)
inputbutton = tk.Button(top, text = 'enter', command=getInput)
inputbutton.pack()


top.mainloop()


import webbrowser
import pyodbc 

conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=mrp1;'
                      'Database=ESIDEMO;'
                      'Trusted_Connection=yes;')

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path

xlsx_file = Path('Documents', inp)
fil = 'Documents'
wb_obj = openpyxl.load_workbook(xlsx_file) 
target = wb_obj['Sheet1']


wb = Workbook()
wb1 = wb.active
print(dnn)

ws = wb_obj.active



conn1 = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=mrp1;'
                      'Database=ESIDEMO;'
                      'Trusted_Connection=yes;')

import pyodbc 

conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=mrp1;'
                      'Database=ESIDEMO;'
                      'Trusted_Connection=yes;')

cursor = conn.cursor()
cursor.execute('SELECT * FROM ESIDEMO.dbo.POFVP')
            

# Read the active sheet:
content = []
#sheet = wb_obj.active

rownum = 0
colnum = 0

count = 0  # 
count1 = 0 #
for row in ws:
    count = 0
    for cell in row:
        count += 1
        if cell.value == 'Manufacturer Part Number':
            count1 = count
            continue
        if count == count1:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM ESIDEMO.dbo.POFVP WHERE POFVP.PART_ID LIKE ' + dnn)
            y = ""
            count2 = 0
            count3 = 0
            rownum +=1
            for row in cursor:
                count2 = 0
                for x in row:
                    count2 += 1
                    if count2 == 2:
                        y = str(x)
                        continue
                    if count2 == 7:
                        if str(x) == str(cell.value):
                            count3 +=1
                            content.append(y)
                            if count3 == 1:
                                wb1.cell(row = rownum,column = count3).value = y
                            else:
                                wb1.cell(row= rownum ,column = count3).value = y


print('Done')
wb.save('C:\\Users\\slee\\Desktop\\data1.xlsx')
